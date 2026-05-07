#!/usr/bin/env python3
"""
Registration API for Make Money AI journal access.
Saves subscriber details (name, email, subscription type) to Excel file.
"""

import os
import json
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed. Excel export will be skipped.")

# Configuration
EXCEL_PATH = r"C:\Users\dansk\Claude\@ team-ai.biz\data\makemoney-subscribers.xlsx"
SUBSCRIBERS_DB = "subscribers.db"

def ensure_excel_file():
    """Create Excel file if it doesn't exist, with headers."""
    if not HAS_OPENPYXL:
        return

    Path(EXCEL_PATH).parent.mkdir(parents=True, exist_ok=True)

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Subscribers"

        # Header row
        headers = ["Timestamp", "Name", "Email", "Daily", "Weekly", "Monthly", "IP Address"]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15

        wb.save(EXCEL_PATH)

def add_to_excel(name, email, subscriptions, ip_address=""):
    """Append subscriber row to Excel file."""
    if not HAS_OPENPYXL:
        return

    try:
        ensure_excel_file()

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # Determine subscription flags
        daily = "Yes" if "daily" in subscriptions else ""
        weekly = "Yes" if "weekly" in subscriptions else ""
        monthly = "Yes" if "monthly" in subscriptions else ""

        # Add row
        row = [
            datetime.now().isoformat(),
            name,
            email,
            daily,
            weekly,
            monthly,
            ip_address
        ]
        ws.append(row)

        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"[ERROR] Failed to write Excel: {e}")

def add_to_sqlite(name, email, subscriptions):
    """Add subscriber to SQLite database (for backwards compatibility)."""
    try:
        conn = sqlite3.connect(SUBSCRIBERS_DB)
        c = conn.cursor()

        # Create table if not exists
        c.execute('''
            CREATE TABLE IF NOT EXISTS subscribers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT,
                name TEXT,
                email TEXT UNIQUE,
                daily INTEGER,
                weekly INTEGER,
                monthly INTEGER
            )
        ''')

        # Insert
        daily_flag = 1 if "daily" in subscriptions else 0
        weekly_flag = 1 if "weekly" in subscriptions else 0
        monthly_flag = 1 if "monthly" in subscriptions else 0

        c.execute(
            '''INSERT OR REPLACE INTO subscribers (timestamp, name, email, daily, weekly, monthly)
               VALUES (?, ?, ?, ?, ?, ?)''',
            (datetime.now().isoformat(), name, email, daily_flag, weekly_flag, monthly_flag)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[ERROR] Failed to write SQLite: {e}")

def handle_registration(request_json, client_ip=""):
    """Handle registration request. Returns (success, message)."""
    try:
        # Validate input
        name = request_json.get("name", "").strip()
        email = request_json.get("email", "").strip()
        subscriptions = request_json.get("subscriptions", [])

        if not name or not email:
            return False, "Name and email are required"

        if "@" not in email:
            return False, "Invalid email format"

        if not subscriptions or len(subscriptions) == 0:
            return False, "At least one subscription option must be selected"

        # Save to Excel and SQLite
        add_to_excel(name, email, subscriptions, client_ip)
        add_to_sqlite(name, email, subscriptions)

        return True, f"Welcome, {name}! Your registration is confirmed."
    except Exception as e:
        return False, f"Registration failed: {str(e)}"
